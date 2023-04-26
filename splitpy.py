import openpyxl

# Carrega a planilha base_bruta
wb = openpyxl.load_workbook('base_bruta.xlsx')
ws_bruta = wb.active

# Cria a planilha base_tratada
wb_tratada = openpyxl.Workbook()
ws_tratada = wb_tratada.active

# Escreve os cabeçalhos na planilha base_tratada
for col in range(1, ws_bruta.max_column + 1):
    ws_tratada.cell(row=1, column=col).value = ws_bruta.cell(row=1, column=col).value

# Percorre a planilha base_bruta e escreve os dados na planilha base_tratada
linha_login = 2
cont_grupo = 2
while ws_bruta.cell(row=linha_login, column=1).value is not None:
    login = ws_bruta.cell(row=linha_login, column=1).value
    grupo_ad = ws_bruta.cell(row=linha_login, column=6).value.split("^")
    nome = ws_bruta.cell(row=linha_login, column=2).value
    custon13 = ws_bruta.cell(row=linha_login, column=8).value
    custon14 = ws_bruta.cell(row=linha_login, column=9).value
    custon15 = ws_bruta.cell(row=linha_login, column=10).value
    matricula_gerente = ws_bruta.cell(row=linha_login, column=11).value

    for grupo in grupo_ad:
        ws_tratada.cell(row=cont_grupo, column=1).value = login
        ws_tratada.cell(row=cont_grupo, column=2).value = nome
        ws_tratada.cell(row=cont_grupo, column=6).value = grupo
        ws_tratada.cell(row=cont_grupo, column=12).value = matricula_gerente
        ws_tratada.cell(row=cont_grupo, column=8).value = custon13
        ws_tratada.cell(row=cont_grupo, column=9).value = login
        ws_tratada.cell(row=cont_grupo, column=10).value = nome
        ws_tratada.cell(row=cont_grupo, column=15).value = custon15
        ws_tratada.cell(row=cont_grupo, column=14).value = custon14
        cont_grupo += 1

    linha_login += 1

# Salva a planilha base_tratada
wb_tratada.save('base_tratada.xlsx')
